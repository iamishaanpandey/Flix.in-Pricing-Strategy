"""
06_excel_export.py
------------------
Builds the complete Excel deliverable from pipeline outputs.

Sheets:
  1. Executive Summary    - KPI dashboard with key metrics
  2. Flagging Output      - Full flagged bus listing (main deliverable)
  3. Peer Group Detail    - Comparable buses per Flixbus row
  4. Logic Explanation    - Written methodology
  5. Automation Plan      - MVP pipeline description
"""

import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import ColorScaleRule, CellIsRule, FormulaRule
import os

OUT = "outputs"

print("Loading pipeline outputs ...")
final   = pd.read_parquet(os.path.join(OUT, "05_final_output.parquet"))
peers   = pd.read_parquet(os.path.join(OUT, "03_peer_groups.parquet"))
summary_df = pd.read_csv(os.path.join(OUT, "05_executive_summary.csv"))

print(f"  Final output: {len(final):,} rows")
print(f"  Peer groups:  {len(peers):,} rows")

# ── Style helpers ─────────────────────────────────────────────
def hex_fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def make_border(style="thin"):
    s = Side(style=style, color="CCCCCC")
    return Border(left=s, right=s, top=s, bottom=s)

def header_style(ws, row, col, value, bg="1A2D42", fg="FFFFFF", size=11, bold=True, center=True):
    c = ws.cell(row=row, column=col, value=value)
    c.font = Font(name="Calibri", bold=bold, color=fg, size=size)
    c.fill = hex_fill(bg)
    c.alignment = Alignment(horizontal="center" if center else "left",
                            vertical="center", wrap_text=True)
    c.border = make_border()
    return c

def data_cell(ws, row, col, value, bold=False, color="000000", bg=None,
              align="left", num_format=None, wrap=False):
    c = ws.cell(row=row, column=col, value=value)
    c.font = Font(name="Calibri", bold=bold, color=color, size=10)
    if bg:
        c.fill = hex_fill(bg)
    c.alignment = Alignment(horizontal=align, vertical="center", wrap_text=wrap)
    c.border = make_border()
    if num_format:
        c.number_format = num_format
    return c

def section_header(ws, row, col, value, width=10, bg="0F1C2E", fg="00C2CB"):
    ws.merge_cells(start_row=row, start_column=col,
                   end_row=row, end_column=col+width-1)
    c = ws.cell(row=row, column=col, value=value)
    c.font = Font(name="Calibri", bold=True, color=fg, size=12)
    c.fill = hex_fill(bg)
    c.alignment = Alignment(horizontal="left", vertical="center")
    return c

# ── Urgency colours ───────────────────────────────────────────
URGENCY_COLOURS = {
    "URGENT":            ("E84545", "FFFFFF"),
    "HIGH":              ("F4A623", "000000"),
    "MONITOR":           ("F4A623", "000000"),
    "OPPORTUNITY":       ("2EC27E", "FFFFFF"),
    "INVESTIGATE":       ("3B82F6", "FFFFFF"),
    "INVESTIGATE_EARLY": ("A78BFA", "FFFFFF"),
    "REVIEW":            ("7B9BAF", "FFFFFF"),
    "CONSIDER_RAISE":    ("00C2CB", "000000"),
    "OPTIMAL":           ("E8F5E9", "2E7D32"),
    "SKIP":              ("F5F5F5", "999999"),
}

FLAG_COLOURS = {
    "OVERPRICED":          ("E84545", "FFFFFF"),
    "OVERPRICED_JUSTIFIED":("F4A623", "000000"),
    "UNDERPRICED":         ("3B82F6", "FFFFFF"),
    "OK":                  ("E8F5E9", "2E7D32"),
    "NO_DATA":             ("F5F5F5", "999999"),
    "NO_PEERS":            ("F5F5F5", "999999"),
}

wb = Workbook()

# ════════════════════════════════════════════════════════════
# SHEET 1 — EXECUTIVE SUMMARY
# ════════════════════════════════════════════════════════════
ws1 = wb.active
ws1.title = "Executive Summary"
ws1.sheet_view.showGridLines = False
ws1.column_dimensions["A"].width = 2

# Title
ws1.merge_cells("B2:K3")
c = ws1["B2"]
c.value = "FlixBus Pricing Intelligence — Executive Summary"
c.font = Font(name="Calibri", bold=True, size=18, color="FFFFFF")
c.fill = hex_fill("0F1C2E")
c.alignment = Alignment(horizontal="left", vertical="center")

ws1.merge_cells("B4:K4")
c = ws1["B4"]
c.value = "Automated pricing anomaly detection across 59 routes | 4 extraction snapshots | 30,888 Flixbus listings analysed"
c.font = Font(name="Calibri", size=10, color="7B9BAF", italic=True)
c.fill = hex_fill("162338")
c.alignment = Alignment(horizontal="left", vertical="center")
ws1.row_dimensions[2].height = 30
ws1.row_dimensions[4].height = 18

# KPI cards — row 6
kpis = [
    ("Total Buses\nAnalysed",    30888,     "0,0",    "1A2D42", "00C2CB"),
    ("🔴 URGENT\nFlags",         1254,      "0,0",    "E84545", "FFFFFF"),
    ("🟡 MONITOR\nFlags",        1025,      "0,0",    "F4A623", "000000"),
    ("💰 OPPORTUNITY\nFlags",    13,        "0,0",    "2EC27E", "FFFFFF"),
    ("📊 Revenue\nImpact Est.",  4857755,   "₹#,##0", "1A2D42", "F4A623"),
]
col = 2
for label, val, fmt, bg, fg in kpis:
    ws1.merge_cells(start_row=6, start_column=col, end_row=6, end_column=col+1)
    ws1.merge_cells(start_row=7, start_column=col, end_row=7, end_column=col+1)
    ws1.merge_cells(start_row=8, start_column=col, end_row=8, end_column=col+1)
    lc = ws1.cell(row=6, column=col, value=label)
    lc.font = Font(name="Calibri", size=9, color="AAAAAA" if bg=="1A2D42" else "000000", bold=True)
    lc.fill = hex_fill(bg)
    lc.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    vc = ws1.cell(row=7, column=col, value=val)
    vc.font = Font(name="Calibri", size=20, bold=True, color=fg)
    vc.fill = hex_fill(bg)
    vc.number_format = fmt
    vc.alignment = Alignment(horizontal="center", vertical="center")
    bc = ws1.cell(row=8, column=col)
    bc.fill = hex_fill(bg)
    col += 2

ws1.row_dimensions[6].height = 28
ws1.row_dimensions[7].height = 36
ws1.row_dimensions[8].height = 8

# Flag breakdown table
section_header(ws1, 10, 2, "  FLAG DISTRIBUTION", width=5)
flag_data = [
    ("OVERPRICED",           1284,  "Overpriced — no quality justification",       "E84545"),
    ("OVERPRICED_JUSTIFIED", 1018,  "Overpriced but higher quality than peers",    "F4A623"),
    ("UNDERPRICED",          4453,  "Priced below statistical lower bound",         "3B82F6"),
    ("OK",                   23470, "Within IQR band — correctly priced",           "2EC27E"),
    ("NO_DATA",              663,   "No comparable peers found",                    "CCCCCC"),
]
hdrs = ["Flag", "Count", "Description", ""]
for i, h in enumerate(hdrs):
    header_style(ws1, 11, 2+i, h, bg="162338", size=10)
for r, (flag, cnt, desc, color) in enumerate(flag_data):
    bg = "F9F9F9" if r % 2 == 0 else "FFFFFF"
    data_cell(ws1, 12+r, 2, flag, bold=True, color=color, bg=bg)
    data_cell(ws1, 12+r, 3, cnt, align="center", bg=bg, num_format="#,##0")
    data_cell(ws1, 12+r, 4, desc, bg=bg)
    ws1.cell(row=12+r, column=5).fill = hex_fill(color)

# Urgency breakdown table
section_header(ws1, 10, 7, "  URGENCY DISTRIBUTION", width=4)
urgency_data = [
    ("URGENT",      1254,  "Overpriced + low occupancy"),
    ("MONITOR",     1025,  "Overpriced + demand exists"),
    ("OPPORTUNITY", 13,    "Underpriced + high demand"),
    ("INVESTIGATE", 4360,  "Underpriced + low load"),
    ("OPTIMAL",     23452, "Correctly priced"),
    ("SKIP",        663,   "Insufficient data"),
]
for i, h in enumerate(["Urgency", "Count", "Meaning"]):
    header_style(ws1, 11, 7+i, h, bg="162338", size=10)
for r, (urg, cnt, meaning) in enumerate(urgency_data):
    bg_col, fg_col = URGENCY_COLOURS.get(urg, ("FFFFFF", "000000"))
    bg = "F9F9F9" if r % 2 == 0 else "FFFFFF"
    data_cell(ws1, 12+r, 7, urg, bold=True, color=bg_col, bg=bg)
    data_cell(ws1, 12+r, 8, cnt, align="center", bg=bg, num_format="#,##0")
    data_cell(ws1, 12+r, 9, meaning, bg=bg)

# Key insight
section_header(ws1, 19, 2, "  KEY INSIGHT", width=9)
ws1.merge_cells("B20:J22")
insight = ws1["B20"]
insight.value = (
    "Flixbus is on average 10.5% CHEAPER than comparable competitors across all routes and snapshots. "
    "This systematic underpricing, combined with 4,360 INVESTIGATE cases (underpriced + low occupancy), "
    "suggests a visibility and discovery problem rather than a pure pricing issue. "
    "1,254 URGENT cases require immediate attention — these buses are overpriced AND have low seat occupancy."
)
insight.font = Font(name="Calibri", size=10, color="D0E8EC")
insight.fill = hex_fill("162338")
insight.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
ws1.row_dimensions[20].height = 50

# Column widths
for col, w in [(2,18),(3,10),(4,38),(5,4),(6,4),(7,16),(8,10),(9,32),(10,4)]:
    ws1.column_dimensions[get_column_letter(col)].width = w

# ════════════════════════════════════════════════════════════
# SHEET 2 — FLAGGING OUTPUT
# ════════════════════════════════════════════════════════════
ws2 = wb.create_sheet("Flagging Output")
ws2.sheet_view.showGridLines = False
ws2.freeze_panes = "A2"

# Title row
ws2.merge_cells("A1:AD1")
c = ws2["A1"]
c.value = "FlixBus Pricing Intelligence — Flagging Output  |  One row per Flixbus bus listing"
c.font = Font(name="Calibri", bold=True, size=12, color="FFFFFF")
c.fill = hex_fill("0F1C2E")
c.alignment = Alignment(horizontal="left", vertical="center")
ws2.row_dimensions[1].height = 22

# Headers
cols_config = [
    # (header, col_key, width, num_format, align)
    ("Route",             "Route_Number",          12, None,      "left"),
    ("Journey Date",      "DOJ",                   14, "DD-MMM-YY","center"),
    ("Day Type",          "Day_Type",              10, None,      "center"),
    ("Bus Category",      "Bus_Category",          16, None,      "left"),
    ("Departure",         "Departure_Time",        10, None,      "center"),
    ("Flix Price (₹)",    "Flix_Price",            13, "#,##0.00","right"),
    ("Peer Median (₹)",   "Peer_Median_Price",     14, "#,##0.00","right"),
    ("Peer Q1 (₹)",       "Peer_Q1_Price",         12, "#,##0.00","right"),
    ("Peer Q3 (₹)",       "Peer_Q3_Price",         12, "#,##0.00","right"),
    ("IQR Lower (₹)",     "IQR_Lower_Bound",       13, "#,##0.00","right"),
    ("IQR Upper (₹)",     "IQR_Upper_Bound",       13, "#,##0.00","right"),
    ("Dev. Abs (₹)",      "Price_Deviation_Abs",   13, "#,##0.00","right"),
    ("Dev. %",            "Price_Deviation_Pct",   10, "0.00%",   "right"),
    ("Stage 1 Flag",      "Stage1_Flag",           14, None,      "center"),
    ("Quality Adj.",      "Quality_Adjustment",    18, None,      "left"),
    ("Final Flag",        "Final_Flag",            18, None,      "center"),
    ("Urgency",           "Urgency_Label",         16, None,      "center"),
    ("Action",            "Urgency_Action",        40, None,      "left"),
    ("Confidence",        "Confidence",            11, None,      "center"),
    ("# Peers",           "Peer_Count",            9,  "#,##0",   "center"),
    ("Window (min)",      "Window_Used_Mins",      12, "#,##0",   "center"),
    ("Flix Score",        "Flix_Bus_Score",        11, "0.00",    "center"),
    ("Peer Avg Score",    "Peer_Avg_Bus_Score",    13, "0.00",    "center"),
    ("Score Diff",        "Flix_vs_Peer_Score_Diff",12,"0.00",   "center"),
    ("Occupancy %",       "Flix_Occupancy_Pct",    12, "0.0%",    "center"),
    ("SRP Rank",          "Flix_SRP_Rank",         10, "#,##0",   "center"),
    ("Revenue Impact",    "Revenue_Impact_Est",    15, "₹#,##0",  "right"),
]

for ci, (hdr, _, width, _, _) in enumerate(cols_config, 1):
    header_style(ws2, 2, ci, hdr, bg="1A2D42", fg="00C2CB", size=10)
    ws2.column_dimensions[get_column_letter(ci)].width = width

ws2.row_dimensions[2].height = 36

# Data rows — write in batches for speed
print("Writing Sheet 2 (Flagging Output) ...")

# Sort: urgent first
sort_order = {"URGENT":0,"HIGH":1,"OPPORTUNITY":2,"MONITOR":3,
              "REVIEW":4,"INVESTIGATE":5,"INVESTIGATE_EARLY":6,
              "CONSIDER_RAISE":7,"OPTIMAL":8,"SKIP":9}
final_sorted = final.copy()
final_sorted["_sort"] = final_sorted["Urgency_Label"].map(sort_order).fillna(99)
final_sorted = final_sorted.sort_values(["_sort", "Revenue_Impact_Est"],
                                         ascending=[True, False]).drop("_sort", axis=1)

for ri, (_, row) in enumerate(final_sorted.iterrows()):
    excel_row = ri + 3
    row_bg = "F9F9F9" if ri % 2 == 0 else "FFFFFF"

    for ci, (_, col_key, _, num_fmt, align) in enumerate(cols_config, 1):
        val = row.get(col_key, "")
        if pd.isna(val):
            val = ""
        # Special formatting for pct column
        if col_key == "Price_Deviation_Pct" and val != "":
            val = val / 100
        c = ws2.cell(row=excel_row, column=ci, value=val)
        c.font = Font(name="Calibri", size=9)
        c.fill = hex_fill(row_bg)
        c.alignment = Alignment(horizontal=align, vertical="center")
        c.border = make_border()
        if num_fmt:
            c.number_format = num_fmt

    # Colour the Final Flag cell
    flag_val = row.get("Final_Flag", "")
    flag_ci = next(i for i, (_, k, *_) in enumerate(cols_config, 1) if k == "Final_Flag")
    fc = ws2.cell(row=excel_row, column=flag_ci)
    if flag_val in FLAG_COLOURS:
        bg_c, fg_c = FLAG_COLOURS[flag_val]
        fc.fill = hex_fill(bg_c)
        fc.font = Font(name="Calibri", size=9, bold=True, color=fg_c)
        fc.alignment = Alignment(horizontal="center", vertical="center")

    # Colour the Urgency cell
    urg_val = row.get("Urgency_Label", "")
    urg_ci = next(i for i, (_, k, *_) in enumerate(cols_config, 1) if k == "Urgency_Label")
    uc = ws2.cell(row=excel_row, column=urg_ci)
    if urg_val in URGENCY_COLOURS:
        bg_c, fg_c = URGENCY_COLOURS[urg_val]
        uc.fill = hex_fill(bg_c)
        uc.font = Font(name="Calibri", size=9, bold=True, color=fg_c)
        uc.alignment = Alignment(horizontal="center", vertical="center")

    if ri % 5000 == 0:
        print(f"  {ri:,} / {len(final_sorted):,} rows written ...")

print(f"  Sheet 2 complete: {len(final_sorted):,} rows")

# ════════════════════════════════════════════════════════════
# SHEET 3 — PEER GROUP DETAIL (sample — top 5000 rows)
# ════════════════════════════════════════════════════════════
ws3 = wb.create_sheet("Peer Group Detail")
ws3.sheet_view.showGridLines = False
ws3.freeze_panes = "A2"

ws3.merge_cells("A1:Q1")
c = ws3["A1"]
c.value = "Peer Group Detail — Comparable buses used for each Flixbus listing  |  Showing top 5,000 rows (sort by Flixbus_Row_ID to group)"
c.font = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
c.fill = hex_fill("0F1C2E")
c.alignment = Alignment(horizontal="left", vertical="center")
ws3.row_dimensions[1].height = 20

peer_cols = [
    ("Flixbus Row ID",    "Flixbus_Row_ID",      14),
    ("Route",             "Route_Number",         12),
    ("Journey Date",      "DOJ",                  14),
    ("Bus Category",      "Bus_Category",         16),
    ("Flix Departure",    "Flix_Departure",       12),
    ("Comp Departure",    "Competitor_Departure", 14),
    ("Gap (min)",         "Departure_Gap_Mins",   10),
    ("Comp Price (₹)",    "Competitor_Price",     14),
    ("Comp Bus Score",    "Competitor_Bus_Score", 13),
    ("Comp Operator",     "Competitor_Operator",  22),
    ("Operator Size",     "Operator_Size",        12),
    ("Same Rating Band",  "Same_Rating_Band",     14),
    ("Same Review Tier",  "Same_Review_Tier",     14),
    ("Soft Score",        "Soft_Score",           10),
    ("Peer Group Size",   "Peer_Group_Size",      13),
    ("Confidence",        "Confidence",           11),
    ("Window Used (min)", "Window_Used_Mins",     14),
]

for ci, (hdr, _, width) in enumerate(peer_cols, 1):
    header_style(ws3, 2, ci, hdr, bg="1A2D42", fg="00C2CB", size=10)
    ws3.column_dimensions[get_column_letter(ci)].width = width

ws3.row_dimensions[2].height = 30

# Write top 5000 peer rows
peers_sample = peers.head(5000)
for ri, (_, row) in enumerate(peers_sample.iterrows()):
    excel_row = ri + 3
    row_bg = "F9F9F9" if ri % 2 == 0 else "FFFFFF"
    for ci, (_, col_key, _) in enumerate(peer_cols, 1):
        val = row.get(col_key, "")
        if pd.isna(val): val = ""
        c = ws3.cell(row=excel_row, column=ci, value=val)
        c.font = Font(name="Calibri", size=9)
        c.fill = hex_fill(row_bg)
        c.alignment = Alignment(horizontal="center" if ci > 1 else "left", vertical="center")
        c.border = make_border()

print("  Sheet 3 complete")

# ════════════════════════════════════════════════════════════
# SHEET 4 — LOGIC EXPLANATION
# ════════════════════════════════════════════════════════════
ws4 = wb.create_sheet("Logic Explanation")
ws4.sheet_view.showGridLines = False
ws4.column_dimensions["A"].width = 2
ws4.column_dimensions["B"].width = 26
ws4.column_dimensions["C"].width = 70
ws4.column_dimensions["D"].width = 26

def write_section(ws, start_row, title, rows, col=2):
    section_header(ws, start_row, col, f"  {title}", width=3)
    for i, (label, content) in enumerate(rows):
        r = start_row + 1 + i
        lc = ws.cell(row=r, column=col, value=label)
        lc.font = Font(name="Calibri", bold=True, size=10, color="00C2CB")
        lc.fill = hex_fill("1A2D42")
        lc.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
        lc.border = make_border()
        cc = ws.cell(row=r, column=col+1, value=content)
        cc.font = Font(name="Calibri", size=10, color="212121")
        cc.fill = hex_fill("F9F9F9" if i % 2 == 0 else "FFFFFF")
        cc.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
        cc.border = make_border()
        ws.row_dimensions[r].height = 48
    return start_row + len(rows) + 2

ws4.merge_cells("B1:D2")
c = ws4["B1"]
c.value = "Logic Explanation — FlixBus Pricing Intelligence System"
c.font = Font(name="Calibri", bold=True, size=16, color="FFFFFF")
c.fill = hex_fill("0F1C2E")
c.alignment = Alignment(horizontal="left", vertical="center")
ws4.row_dimensions[1].height = 30

row = 4
row = write_section(ws4, row, "1. WHAT IS ONE ROW?", [
    ("Definition", "Each row in the Flagging Output sheet represents one Flixbus bus listing on a specific route, journey date (DOJ), and extraction snapshot date. The same physical bus appears up to 4 times — once per extraction date — allowing trend analysis across time."),
    ("Dataset scale", "862,388 raw rows → 821,245 after cleaning. 30,888 Flixbus rows across 59 routes, 58 journey dates, and 4 extraction snapshots (Feb 17, Feb 19, Mar 05, Mar 16)."),
    ("Price field used", "Weighted Average Price — a single float representing the weighted average across all fare tiers offered by that bus. Used as the primary comparison metric throughout."),
])

row = write_section(ws4, row, "2. HOW WE DEFINE SIMILAR BUSES (Hierarchical Filter Model)", [
    ("Hard Filter #1 — Route", "Exact Route Number match. A bus on Route 1 is NEVER compared to a bus on Route 2. Non-negotiable."),
    ("Hard Filter #2 — Snapshot Date", "Same Date of Extraction. A Feb 17 price is only compared to other Feb 17 prices. This prevents cross-snapshot dynamic pricing distortions. Non-negotiable."),
    ("Hard Filter #3 — Day Type", "Weekday (Mon–Thu) vs Weekend (Fri–Sun). Friday treated as weekend due to demand patterns. Non-negotiable."),
    ("Hard Filter #4 — Bus Category", "AC/Non-AC + Seater/Sleeper/Semi-Sleeper must match exactly. An AC Sleeper is never compared to a Non-AC Seater. Non-negotiable."),
    ("Hard Filter #5 — Departure Window", "±90 minutes around Flixbus departure time. If fewer than 3 peers found, window relaxes to ±150 minutes. A flag from a <3 peer group is marked Low Confidence."),
    ("Soft Score #1 — Rating Band", "Buses within ±0.5 star band of Flixbus get a soft similarity score boost (+1). Does not exclude peers, only weights them."),
    ("Soft Score #2 — Review Volume Tier", "Tier 1 (500+ reviews), Tier 2 (51-500), Tier 3 (<50). Same tier = +1 soft score. Used as confidence weight."),
    ("Soft Score #3 — Operator Size", "Large / Medium / Small by listing count percentile. Informational context only — shown in output but not used for exclusion."),
])

row = write_section(ws4, row, "3. FLAGGING LOGIC", [
    ("Stage 1 — IQR Statistical Flag", "Compute peer group Median and IQR (Q3-Q1). Lower bound = Median - (1.5 × IQR). Upper bound = Median + (1.5 × IQR). OVERPRICED if Flix price > Upper. UNDERPRICED if < Lower. OK if within bounds. Same logic as box-plot outlier detection — defensible to non-technical audience."),
    ("Stage 2 — Quality Adjustment", "Check if price deviation is explained by Bus Score difference. If Flix Bus Score > Peer Avg by >0.3 AND overpriced → Quality Justified (downgrade urgency). This prevents flagging a genuinely premium product as overpriced. If no quality difference → flag stands at full strength (Not Justified)."),
    ("Urgency Scoring", "Cross-reference flag with Occupancy %. OVERPRICED + low occupancy (<50%) = URGENT. OVERPRICED + high occupancy (>75%) = MONITOR. UNDERPRICED + high occupancy = OPPORTUNITY. UNDERPRICED + low occupancy = INVESTIGATE (check rank/quality/timing, not price)."),
    ("Confidence levels", "High = 5+ peers. Medium = 3-4 peers. Low = <3 peers. Low confidence flags are included but clearly labelled — do not act on them without manual verification."),
])

row = write_section(ws4, row, "4. ASSUMPTIONS & LIMITATIONS", [
    ("Dynamic pricing", "Prices change throughout the day and as departure approaches. Each snapshot is a point-in-time comparison. Flags should be interpreted in the context of when the data was scraped."),
    ("Days to departure range", "Dataset covers 0-30 days before departure. No long-range pricing data available. Occupancy signals close to departure are more reliable than those far out."),
    ("Bus Score comparability", "Bus Score is a platform-level composite metric. Assumed comparable across operators, but may reflect different weighting schemes per operator type."),
    ("AI tool usage", "Claude (Anthropic) was used for: pipeline architecture design, Python script generation, logic framework development, and Excel deliverable construction. All analytical decisions and threshold choices were made by the analyst based on the data."),
])

print("  Sheet 4 complete")

# ════════════════════════════════════════════════════════════
# SHEET 5 — AUTOMATION PLAN
# ════════════════════════════════════════════════════════════
ws5 = wb.create_sheet("Automation Plan")
ws5.sheet_view.showGridLines = False
ws5.column_dimensions["A"].width = 2
ws5.column_dimensions["B"].width = 24
ws5.column_dimensions["C"].width = 72

ws5.merge_cells("B1:C2")
c = ws5["B1"]
c.value = "Automation Plan — MVP Pipeline Architecture"
c.font = Font(name="Calibri", bold=True, size=16, color="FFFFFF")
c.fill = hex_fill("0F1C2E")
c.alignment = Alignment(horizontal="left", vertical="center")
ws5.row_dimensions[1].height = 30

row = 4
row = write_section(ws5, row, "SYSTEM OVERVIEW", [
    ("Architecture", "CSV drop → Power Automate (trigger) → Python scripts (heavy processing) → Output CSVs → Power BI Dataflows (light transforms) → Published Power BI Report (daily auto-refresh). Each layer has a single responsibility — failures in one layer do not affect others."),
    ("Repeatability", "Dropping a new data.xlsx into the SharePoint folder automatically triggers the entire pipeline. The pricing team receives an email when the report is ready. No manual steps required after initial setup."),
], col=2)

row = write_section(ws5, row, "PIPELINE STEPS", [
    ("Step 1 — Data Drop", "New CSV/Excel placed in SharePoint /FlixBus/Data/Raw/. Naming convention: pricing_YYYY-MM-DD.csv. Power Automate watches this folder via file-created trigger."),
    ("Step 2 — Power Automate", "On file creation: (1) Send file path to Python via HTTP call. (2) Await completion. (3) Archive raw file to /Processed/. (4) Trigger Power BI dataset refresh via REST API. (5) Send email notification to pricing team."),
    ("Step 3 — 01_ingest.py", "Load file, validate schema, enforce types, fill NAs with business-logic defaults, drop zero-variance columns, remove true duplicates. Output: 01_validated.parquet."),
    ("Step 4 — 02_features.py", "Engineer derived columns: Day_Type, Bus_Category, Occupancy_Pct, Departure_Mins, Fare_Min_Price, Rating_Band, Review_Tier, Operator_Size, Days_To_Departure. Output: 02_featured.parquet."),
    ("Step 5 — 03_similarity.py", "Vectorised peer grouping via pd.merge on hard filters. Departure window filter. Soft scoring. Peer statistics via groupby. Output: 03_peer_groups.parquet, 03_similarity_summary.parquet."),
    ("Step 6 — 04_flag.py", "Stage 1: IQR statistical flag. Stage 2: Quality adjustment. Outputs flag direction and magnitude per Flixbus row. Output: 04_flags.parquet."),
    ("Step 7 — 05_urgency.py", "Cross-reference flags with occupancy. Assign urgency label and score. Estimate revenue impact. Output: 05_final_output.parquet — feeds Power BI directly."),
    ("Step 8 — Power BI", "Dataflows read final_output.parquet from SharePoint. Power Query handles light transforms (data types, buckets, colour codes). DAX measures compute KPIs. Published report auto-refreshes on dataset trigger."),
], col=2)

row = write_section(ws5, row, "TECH STACK", [
    ("Backend", "Python 3.10+ | pandas | numpy | pyarrow | openpyxl | pyyaml"),
    ("Orchestration", "Microsoft Power Automate — file trigger, HTTP action, Power BI refresh action, email notification"),
    ("Storage", "SharePoint / OneDrive — raw data landing, processed outputs, Power BI data source"),
    ("BI Layer", "Power BI Service — Dataflows (Power Query / M), Dataset (DAX), Published Report (5 pages)"),
    ("Configuration", "config.yaml — all thresholds (IQR multiplier, peer group minimum, departure window, occupancy thresholds) parameterised. Change thresholds without touching code."),
], col=2)

row = write_section(ws5, row, "PHASE ROADMAP", [
    ("Phase 1 — Now (Assignment)", "Python scripts running locally. Excel deliverable generated manually. Power BI loaded from CSV manually. Proves the logic works end-to-end."),
    ("Phase 2 — MVP (Week 1-2)", "Scripts hosted on Azure Function or VM. Power Automate flow live. Full Power BI report published. Daily automated refresh."),
    ("Phase 3 — Production (Month 1+)", "SQL database replaces CSV files (PostgreSQL / Azure SQL). Historical data retained — trend analysis across months. Quarterly threshold recalibration. Operator-level pricing strategy detection."),
], col=2)

print("  Sheet 5 complete")

# ════════════════════════════════════════════════════════════
# SAVE
# ════════════════════════════════════════════════════════════
out_path = os.path.join(OUT, "FlixBus_Pricing_Intelligence.xlsx")
print(f"\nSaving workbook to {out_path} ...")
wb.save(out_path)
print(f"✅ Excel workbook saved: {out_path}")
print(f"   Sheets: Executive Summary | Flagging Output | Peer Group Detail | Logic Explanation | Automation Plan")